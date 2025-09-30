from rest_framework import viewsets, status, permissions, filters
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Q, Count, Avg, Sum
from django.utils import timezone
from django.contrib.auth import get_user_model
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

from .models import Project, ProjectComment, ProjectAttachment, Task, TaskComment, TaskAttachment, TimeEntry, ProjectNote

User = get_user_model()
from .serializers import (
    ProjectSerializer, ProjectListSerializer, ProjectCreateUpdateSerializer,
    ProjectCommentSerializer, ProjectAttachmentSerializer,
    TaskSerializer, TaskListSerializer, TaskCreateUpdateSerializer,
    TaskCommentSerializer, TaskAttachmentSerializer, TimeEntrySerializer,
    ProjectNoteSerializer
)
from .filters import ProjectFilter, TaskFilter
from .permissions import IsProjectManagerOrReadOnly, IsProjectManager, CanViewProject, CanModifyProject


class ProjectViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing projects
    - All authenticated users: Read access
    - PROJECT_MANAGER, manager, admin: Full access (CRUD)
    """
    queryset = Project.objects.all()
    permission_classes = [permissions.IsAuthenticated, CanModifyProject]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = ProjectFilter
    search_fields = ['name', 'description', 'tags']
    ordering_fields = ['name', 'created_at', 'deadline', 'priority', 'status']
    ordering = ['-created_at']
    
    def get_serializer_class(self):
        if self.action == 'list':
            return ProjectListSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            # Use create/update serializer for validation on input,
            # but we'll override response to return full ProjectSerializer
            return ProjectCreateUpdateSerializer
        return ProjectSerializer
    
    def get_queryset(self):
        """
        Filter projects based on user permissions
        """
        user = self.request.user
        
        # For development: allow all authenticated users to see all projects
        if user.is_authenticated:
            return Project.objects.all()
        
        # Admin can see all projects
        if user.role == 'admin':
            return Project.objects.all()
        
        # Manager can see projects they manage or are part of
        if user.role == 'manager':
            return Project.objects.filter(
                Q(manager=user) | Q(team=user)
            ).distinct()
        
        # Other users can only see projects they're part of
        return Project.objects.filter(team=user)
    
    @action(detail=True, methods=['patch'], permission_classes=[permissions.IsAuthenticated])
    def update_progress(self, request, pk=None):
        """
        Update project progress
        """
        try:
            project = self.get_object()
            
            # Check if user has permission to update progress
            user = request.user
            user_role = getattr(user, 'role', None)
            
            # Allow progress updates for team members, managers, and admins
            can_update = (
                user.is_superuser or 
                user.is_staff or 
                user_role == 'admin' or 
                user_role == 'manager' or 
                user_role == 'PROJECT_MANAGER' or
                project.team.filter(id=user.id).exists() or
                project.manager == user
            )
            
            if not can_update:
                return Response({
                    'success': False,
                    'error': 'Permission insuffisante',
                    'message': 'Vous devez être membre de l\'équipe du projet, gestionnaire ou administrateur pour modifier le progrès.',
                    'required_roles': ['admin', 'manager', 'PROJECT_MANAGER', 'membre_équipe']
                }, status=status.HTTP_403_FORBIDDEN)
            
            progress = request.data.get('progress')
            
            if progress is None:
                return Response(
                    {'error': 'Progress value is required'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Validate progress value
            try:
                progress = int(progress)
                if progress < 0 or progress > 100:
                    return Response(
                        {'error': 'Progress must be between 0 and 100'}, 
                        status=status.HTTP_400_BAD_REQUEST
                    )
            except (ValueError, TypeError):
                return Response(
                    {'error': 'Progress must be a valid number'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Update progress
            project.progress = progress
            
            # Auto-update status based on progress
            if progress >= 100:
                if project.status != 'Terminé':
                    project.status = 'Terminé'
                    project.completed_date = timezone.now().date()
            elif progress > 0:
                if project.status == 'Planification':
                    project.status = 'En cours'
                elif project.status == 'Terminé':
                    # If going back from 100% to less than 100%, change to 'En cours'
                    project.status = 'En cours'
                    project.completed_date = None  # Clear completed date
            else:  # progress == 0
                if project.status in ['En cours', 'Terminé']:
                    # If going back to 0%, change to 'Planification'
                    project.status = 'Planification'
                    project.completed_date = None  # Clear completed date
            
            project.save()
            
            # Return updated project data
            serializer = ProjectSerializer(project)
            return Response({
                'success': True,
                'data': serializer.data,
                'message': 'Progress updated successfully'
            })
            
        except Project.DoesNotExist:
            return Response(
                {'error': 'Project not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def perform_create(self, serializer):
        """Set the manager to the current user if not specified"""
        if not serializer.validated_data.get('manager'):
            serializer.save(manager=self.request.user)
        else:
            serializer.save()

    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        # Re-serialize with full serializer so client immediately gets manager_name, etc.
        if response.status_code in [status.HTTP_200_OK, status.HTTP_201_CREATED]:
            project_id = response.data.get('id') if isinstance(response.data, dict) else None
            if project_id:
                instance = Project.objects.get(id=project_id)
                data = ProjectSerializer(instance).data
                return Response(data, status=response.status_code)
        return response

    def update(self, request, *args, **kwargs):
        response = super().update(request, *args, **kwargs)
        if response.status_code in [status.HTTP_200_OK, status.HTTP_201_CREATED]:
            instance = self.get_object()
            data = ProjectSerializer(instance).data
            return Response(data, status=response.status_code)
        return response

    def partial_update(self, request, *args, **kwargs):
        # Get the project instance before update
        instance = self.get_object()
        old_status = instance.status
        old_progress = instance.progress
        
        # Check if status is being updated
        new_status = request.data.get('status')
        
        response = super().partial_update(request, *args, **kwargs)
        if response.status_code in [status.HTTP_200_OK, status.HTTP_201_CREATED]:
            # Get updated instance
            instance = self.get_object()
            
            # Auto-update progress based on status change
            if new_status and new_status != old_status:
                if new_status == 'Terminé' and old_progress < 100:
                    instance.progress = 100
                    instance.completed_date = timezone.now().date()
                elif new_status == 'Planification' and old_progress > 0:
                    instance.progress = 0
                    instance.completed_date = None
                elif new_status == 'En cours' and old_progress == 0:
                    instance.progress = 50  # Set to middle value when starting
                elif new_status in ['En attente', 'En retard'] and old_progress == 100:
                    instance.progress = 75  # Set to high value when pausing
                
                instance.save()
            
            data = ProjectSerializer(instance).data
            return Response(data, status=response.status_code)
        return response
    
    @action(detail=True, methods=['get'])
    def statistics(self, request, pk=None):
        """Get project statistics"""
        project = self.get_object()
        
        # Check permissions
        if not self._has_project_access(project, request.user):
            return Response(
                {'error': 'Permission denied'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        tasks = project.tasks.all()
        completed_tasks = tasks.filter(status='completed')
        overdue_tasks = tasks.filter(
            due_date__lt=timezone.now(),
            status__in=['not_started', 'in_progress']
        )
        
        stats = {
            'total_tasks': tasks.count(),
            'completed_tasks': completed_tasks.count(),
            'in_progress_tasks': tasks.filter(status='in_progress').count(),
            'overdue_tasks': overdue_tasks.count(),
            'completion_rate': round(
                (completed_tasks.count() / tasks.count() * 100) if tasks.count() > 0 else 0, 2
            ),
            'average_task_duration': round(
                tasks.aggregate(avg=Avg('actual_time'))['avg'] or 0, 2
            ),
            'budget_utilization': project.budget_utilization,
            'team_size': project.team_count,
            'days_remaining': (project.deadline - timezone.now().date()).days if project.deadline else None
        }
        
        return Response(stats)
    
    @action(detail=True, methods=['post'])
    def add_team_member(self, request, pk=None):
        """Add a team member to the project"""
        project = self.get_object()
        
        # Check permissions
        if not self._can_manage_project(project, request.user):
            return Response({
                'success': False,
                'error': 'Permission insuffisante',
                'message': 'Vous devez être gestionnaire du projet ou administrateur pour ajouter des membres à l\'équipe.',
                'required_roles': ['admin', 'manager', 'PROJECT_MANAGER']
            }, status=status.HTTP_403_FORBIDDEN)
        
        user_id = request.data.get('user_id')
        if not user_id:
            return Response(
                {'error': 'user_id is required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            from authentication.models import User
            user = User.objects.get(id=user_id)
            project.team.add(user)
            return Response({'message': 'Team member added successfully'})
        except User.DoesNotExist:
            return Response(
                {'error': 'User not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )
    
    @action(detail=True, methods=['delete'])
    def remove_team_member(self, request, pk=None):
        """Remove a team member from the project"""
        project = self.get_object()
        
        # Check permissions
        if not self._can_manage_project(project, request.user):
            return Response(
                {'error': 'Permission denied'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        user_id = request.data.get('user_id')
        if not user_id:
            return Response(
                {'error': 'user_id is required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            from authentication.models import User
            user = User.objects.get(id=user_id)
            project.team.remove(user)
            return Response({'message': 'Team member removed successfully'})
        except User.DoesNotExist:
            return Response(
                {'error': 'User not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )
    
    def _has_project_access(self, project, user):
        """Check if user has access to the project"""
        return (
            user.role == 'admin' or 
            project.manager == user or 
            project.team.filter(id=user.id).exists()
        )
    
    def _can_manage_project(self, project, user):
        """Check if user can manage the project"""
        return user.role == 'admin' or project.manager == user


class ProjectCommentListCreateView(viewsets.ModelViewSet):
    """
    ViewSet for project comments
    """
    serializer_class = ProjectCommentSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.OrderingFilter]
    ordering = ['-created_at']
    
    def get_queryset(self):
        project_id = self.kwargs.get('project_id')
        return ProjectComment.objects.filter(project_id=project_id)
    
    def perform_create(self, serializer):
        project_id = self.kwargs.get('project_id')
        serializer.save(project_id=project_id)


class ProjectAttachmentListCreateView(viewsets.ModelViewSet):
    """
    ViewSet for project attachments
    """
    serializer_class = ProjectAttachmentSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.OrderingFilter]
    ordering = ['-uploaded_at']
    
    def get_queryset(self):
        project_id = self.kwargs.get('project_id')
        return ProjectAttachment.objects.filter(project_id=project_id)
    
    def perform_create(self, serializer):
        project_id = self.kwargs.get('project_id')
        serializer.save(project_id=project_id)


class ProjectStatisticsView(viewsets.ViewSet):
    """
    ViewSet for project statistics
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def list(self, request):
        """Get overall project statistics"""
        user = request.user
        
        # Filter projects based on user permissions
        if user.role == 'admin':
            projects = Project.objects.all()
        elif user.role == 'manager':
            projects = Project.objects.filter(
                Q(manager=user) | Q(team=user)
            ).distinct()
        else:
            # For all other roles, show all projects (read-only access)
            projects = Project.objects.all()
        
        total_projects = projects.count()
        active_projects = projects.filter(status='in_progress').count()
        completed_projects = projects.filter(status='completed').count()
        overdue_projects = projects.filter(
            deadline__lt=timezone.now().date(),
            status__in=['planning', 'in_progress']
        ).count()
        
        stats = {
            'total_projects': total_projects,
            'active_projects': active_projects,
            'completed_projects': completed_projects,
            'overdue_projects': overdue_projects,
            'completion_rate': round(
                (completed_projects / total_projects * 100) if total_projects > 0 else 0, 2
            ),
            'projects_by_status': dict(projects.values('status').annotate(count=Count('id')).values_list('status', 'count')),
            'projects_by_priority': dict(projects.values('priority').annotate(count=Count('id')).values_list('priority', 'count')),
            'projects_by_category': dict(projects.values('category').annotate(count=Count('id')).values_list('category', 'count'))
        }
        
        return Response(stats)


class TaskViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing tasks
    - All authenticated users: Read access
    - PROJECT_MANAGER, manager, admin: Full access (CRUD)
    """
    queryset = Task.objects.all()
    permission_classes = [permissions.IsAuthenticated, CanModifyProject]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = TaskFilter
    search_fields = ['title', 'description', 'tags']
    ordering_fields = ['title', 'created_at', 'due_date', 'priority', 'status']
    ordering = ['-created_at']
    
    def get_serializer_class(self):
        if self.action == 'list':
            return TaskListSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return TaskCreateUpdateSerializer
        return TaskSerializer
    
    def get_queryset(self):
        """
        Filter tasks based on user permissions
        """
        user = self.request.user
        
        # For development: allow all authenticated users to see all tasks
        if user.is_authenticated:
            return Task.objects.all()
        
        # Admin can see all tasks
        if user.role == 'admin':
            return Task.objects.all()
        
        # Manager can see tasks from projects they manage or are part of
        if user.role == 'manager':
            return Task.objects.filter(
                Q(project__manager=user) | Q(project__team=user)
            ).distinct()
        
        # Other users can only see tasks assigned to them or from projects they're part of
        return Task.objects.filter(
            Q(assignee=user) | Q(project__team=user)
        ).distinct()
    
    def perform_create(self, serializer):
        """Set the reporter to the current user"""
        serializer.save(reporter=self.request.user)
    
    @action(detail=True, methods=['post'])
    def assign(self, request, pk=None):
        """Assign task to a user"""
        task = self.get_object()
        
        # Check permissions
        if not self._can_manage_task(task, request.user):
            return Response({
                'success': False,
                'error': 'Permission insuffisante',
                'message': 'Vous devez être gestionnaire du projet ou administrateur pour assigner des tâches.',
                'required_roles': ['admin', 'manager', 'PROJECT_MANAGER']
            }, status=status.HTTP_403_FORBIDDEN)
        
        user_id = request.data.get('user_id')
        if not user_id:
            return Response(
                {'error': 'user_id is required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            from authentication.models import User
            user = User.objects.get(id=user_id)
            task.assignee = user
            task.save()
            return Response({'message': 'Task assigned successfully'})
        except User.DoesNotExist:
            return Response(
                {'error': 'User not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )
    
    @action(detail=True, methods=['post'])
    def change_status(self, request, pk=None):
        """Change task status"""
        task = self.get_object()
        
        # Check permissions
        if not self._can_manage_task(task, request.user):
            return Response({
                'success': False,
                'error': 'Permission insuffisante',
                'message': 'Vous devez être assigné à cette tâche, gestionnaire du projet ou administrateur pour changer le statut.',
                'required_roles': ['admin', 'manager', 'PROJECT_MANAGER', 'assigné_à_la_tâche']
            }, status=status.HTTP_403_FORBIDDEN)
        
        new_status = request.data.get('status')
        if not new_status:
            return Response(
                {'error': 'status is required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if new_status not in [choice[0] for choice in Task.STATUS_CHOICES]:
            return Response(
                {'error': 'Invalid status'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        task.status = new_status
        if new_status == 'completed':
            task.completed_date = timezone.now()
        task.save()
        
        return Response({'message': 'Task status updated successfully'})
    
    @action(detail=True, methods=['get'])
    def time_entries(self, request, pk=None):
        """Get time entries for a task"""
        task = self.get_object()
        
        # Check permissions
        if not self._has_task_access(task, request.user):
            return Response(
                {'error': 'Permission denied'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        time_entries = task.time_entries.all()
        serializer = TimeEntrySerializer(time_entries, many=True)
        return Response(serializer.data)
    
    def _has_task_access(self, task, user):
        """Check if user has access to the task"""
        return (
            user.role == 'admin' or 
            task.assignee == user or 
            task.reporter == user or
            task.project.manager == user or
            task.project.team.filter(id=user.id).exists()
        )
    
    def _can_manage_task(self, task, user):
        """Check if user can manage the task"""
        return (
            user.role == 'admin' or 
            task.project.manager == user or
            task.assignee == user
        )


class TaskCommentListCreateView(viewsets.ModelViewSet):
    """
    ViewSet for task comments
    """
    serializer_class = TaskCommentSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.OrderingFilter]
    ordering = ['-created_at']
    
    def get_queryset(self):
        task_id = self.kwargs.get('task_id')
        return TaskComment.objects.filter(task_id=task_id)
    
    def perform_create(self, serializer):
        task_id = self.kwargs.get('task_id')
        serializer.save(task_id=task_id)


class TaskAttachmentListCreateView(viewsets.ModelViewSet):
    """
    ViewSet for task attachments
    """
    serializer_class = TaskAttachmentSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.OrderingFilter]
    ordering = ['-uploaded_at']
    
    def get_queryset(self):
        task_id = self.kwargs.get('task_id')
        return TaskAttachment.objects.filter(task_id=task_id)
    
    def perform_create(self, serializer):
        task_id = self.kwargs.get('task_id')
        serializer.save(task_id=task_id)


class TimeEntryListCreateView(viewsets.ModelViewSet):
    """
    ViewSet for time entries
    """
    serializer_class = TimeEntrySerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.OrderingFilter]
    ordering = ['-date', '-created_at']
    
    def get_queryset(self):
        task_id = self.kwargs.get('task_id')
        return TimeEntry.objects.filter(task_id=task_id)
    
    def perform_create(self, serializer):
        task_id = self.kwargs.get('task_id')
        serializer.save(task_id=task_id)


class TaskStatisticsView(viewsets.ViewSet):
    """
    ViewSet for task statistics
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def list(self, request):
        """Get overall task statistics"""
        user = request.user
        
        # Filter tasks based on user permissions
        if user.role == 'admin':
            tasks = Task.objects.all()
        elif user.role == 'manager':
            tasks = Task.objects.filter(
                Q(project__manager=user) | Q(project__team=user)
            ).distinct()
        else:
            # For all other roles, show all tasks (read-only access)
            tasks = Task.objects.all()
        
        total_tasks = tasks.count()
        completed_tasks = tasks.filter(status='completed').count()
        in_progress_tasks = tasks.filter(status='in_progress').count()
        overdue_tasks = tasks.filter(
            due_date__lt=timezone.now(),
            status__in=['not_started', 'in_progress']
        ).count()
        
        # Time tracking statistics
        total_estimated_time = tasks.aggregate(total=Sum('estimated_time'))['total'] or 0
        total_actual_time = tasks.aggregate(total=Sum('actual_time'))['total'] or 0
        
        stats = {
            'total_tasks': total_tasks,
            'completed_tasks': completed_tasks,
            'in_progress_tasks': in_progress_tasks,
            'overdue_tasks': overdue_tasks,
            'completion_rate': round(
                (completed_tasks / total_tasks * 100) if total_tasks > 0 else 0, 2
            ),
            'total_estimated_time': float(total_estimated_time),
            'total_actual_time': float(total_actual_time),
            'time_variance': float(total_actual_time - total_estimated_time),
            'tasks_by_status': dict(tasks.values('status').annotate(count=Count('id')).values_list('status', 'count')),
            'tasks_by_priority': dict(tasks.values('priority').annotate(count=Count('id')).values_list('priority', 'count')),
            'tasks_by_type': dict(tasks.values('task_type').annotate(count=Count('id')).values_list('task_type', 'count'))
        }
        
        return Response(stats)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def dashboard_data(request):
    """
    Get comprehensive dashboard data for the authenticated user
    """
    user = request.user
    
    try:
        # Filter data based on user permissions
        user_role = getattr(user, 'role', 'user')
        if user_role == 'admin':
            projects = Project.objects.all()
            tasks = Task.objects.all()
        elif user_role == 'manager':
            projects = Project.objects.filter(
                Q(manager=user) | Q(team=user)
            ).distinct()
            tasks = Task.objects.filter(
                Q(project__manager=user) | Q(project__team=user)
            ).distinct()
        else:
            # For all other roles, show all data (read-only access)
            projects = Project.objects.all()
            tasks = Task.objects.all()
        
        # Project statistics
        total_projects = projects.count()
        active_projects = projects.filter(status='in_progress').count()
        completed_projects = projects.filter(status='completed').count()
        planning_projects = projects.filter(status='planning').count()
        
        # Task statistics
        total_tasks = tasks.count()
        completed_tasks = tasks.filter(status='completed').count()
        in_progress_tasks = tasks.filter(status='in_progress').count()
        not_started_tasks = tasks.filter(status='not_started').count()
        overdue_tasks = tasks.filter(
            due_date__lt=timezone.now(),
            status__in=['not_started', 'in_progress']
        ).count()
        
        # Recent projects (last 3)
        recent_projects = projects.order_by('-created_at')[:3]
        recent_projects_data = []
        for project in recent_projects:
            # Calculate progress based on completed tasks
            total_tasks = project.tasks.count()
            completed_tasks = project.tasks.filter(status='completed').count()
            progress = round((completed_tasks / total_tasks * 100) if total_tasks > 0 else 0, 1)
            
            recent_projects_data.append({
                'id': project.id,
                'name': project.name,
                'description': getattr(project, 'description', 'Aucune description disponible'),
                'status': project.status,
                'priority': project.priority,
                'progress': progress,
                'deadline': project.deadline,
                'manager_name': project.manager.full_name if project.manager else 'Non assigné',
                'team_count': project.team.count(),
                'tasks_count': total_tasks,
                'completed_tasks_count': completed_tasks,
                'budget': float(getattr(project, 'budget', 0)),
                'spent': float(getattr(project, 'spent', 0)),
                'created_at': project.created_at,
                'updated_at': project.updated_at
            })
        
        # Upcoming tasks (next 5)
        upcoming_tasks = tasks.filter(
            due_date__gte=timezone.now(),
            status__in=['not_started', 'in_progress']
        ).order_by('due_date')[:5]
        upcoming_tasks_data = []
        for task in upcoming_tasks:
            upcoming_tasks_data.append({
                'id': task.id,
                'title': task.title,
                'status': task.status,
                'priority': task.priority,
                'due_date': task.due_date,
                'project_name': task.project.name if task.project else 'Projet supprimé',
                'assignee_name': task.assignee.full_name if task.assignee else 'Non assigné',
                'estimated_time': float(getattr(task, 'estimated_time', 0)),
                'actual_time': float(getattr(task, 'actual_time', 0))
            })
        
        # My tasks (for current user)
        my_tasks = tasks.filter(assignee=user).order_by('-created_at')[:5]
        my_tasks_data = []
        for task in my_tasks:
            my_tasks_data.append({
                'id': task.id,
                'title': task.title,
                'status': task.status,
                'priority': task.priority,
                'due_date': task.due_date,
                'project_name': task.project.name if task.project else 'Projet supprimé',
                'estimated_time': float(getattr(task, 'estimated_time', 0)),
                'actual_time': float(getattr(task, 'actual_time', 0))
            })
        
        # Team members (if user is manager or admin)
        team_members = []
        if hasattr(user, 'role') and user.role in ['admin', 'manager']:
            try:
                team_members_queryset = User.objects.filter(
                    Q(managed_projects__manager=user) | 
                    Q(projects__manager=user)
                ).distinct()[:10]
                team_members_data = []
                for member in team_members_queryset:
                    member_tasks = tasks.filter(assignee=member)
                    team_members_data.append({
                        'id': member.id,
                        'username': member.username,
                        'full_name': getattr(member, 'full_name', member.username),
                        'role': getattr(member, 'role', 'user'),
                        'department': getattr(member, 'department', ''),
                        'assigned_tasks': member_tasks.count(),
                        'completed_tasks': member_tasks.filter(status='completed').count()
                    })
                team_members = team_members_data
            except Exception as e:
                # If there's an issue with team members, just set empty list
                team_members = []
        
        # Performance metrics
        total_estimated_time = tasks.aggregate(total=Sum('estimated_time'))['total'] or 0
        total_actual_time = tasks.aggregate(total=Sum('actual_time'))['total'] or 0
        
        # Budget utilization
        total_budget = projects.aggregate(total=Sum('budget'))['total'] or 0
        total_spent = projects.aggregate(total=Sum('spent'))['total'] or 0
        
        dashboard_data = {
            'user': {
                'id': user.id,
                'username': user.username,
                'full_name': getattr(user, 'full_name', user.username),
                'role': getattr(user, 'role', 'user'),
                'department': getattr(user, 'department', '')
            },
            'statistics': {
                'projects': {
                    'total': total_projects,
                    'active': active_projects,
                    'completed': completed_projects,
                    'planning': planning_projects,
                    'completion_rate': round((completed_projects / total_projects * 100) if total_projects > 0 else 0, 2)
                },
                'tasks': {
                    'total': total_tasks,
                    'completed': completed_tasks,
                    'in_progress': in_progress_tasks,
                    'not_started': not_started_tasks,
                    'overdue': overdue_tasks,
                    'completion_rate': round((completed_tasks / total_tasks * 100) if total_tasks > 0 else 0, 2)
                },
                'time_tracking': {
                    'total_estimated': float(total_estimated_time),
                    'total_actual': float(total_actual_time),
                    'variance': float(total_actual_time - total_estimated_time),
                    'efficiency': round((total_estimated_time / total_actual_time * 100) if total_actual_time > 0 else 0, 2)
                },
                'budget': {
                    'total_budget': float(total_budget),
                    'total_spent': float(total_spent),
                    'utilization_rate': round((total_spent / total_budget * 100) if total_budget > 0 else 0, 2)
                }
            },
            'recent_projects': recent_projects_data,
            'upcoming_tasks': upcoming_tasks_data,
            'my_tasks': my_tasks_data,
            'team_members': team_members,
            'timestamp': timezone.now().isoformat()
        }
        
        return Response({
            'success': True,
            'data': dashboard_data
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ProjectNoteViewSet(viewsets.ModelViewSet):
    """
    ViewSet for project notes (social media style comments)
    """
    serializer_class = ProjectNoteSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        """Filter notes by project if project_id is provided"""
        queryset = ProjectNote.objects.all()
        project_id = self.request.query_params.get('project_id')
        if project_id:
            queryset = queryset.filter(project_id=project_id)
        return queryset
    
    def create(self, request, *args, **kwargs):
        """Create a new note"""
        serializer = self.get_serializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response({
            'success': True,
            'data': serializer.data,
            'message': 'Note created successfully'
        }, status=status.HTTP_201_CREATED)
    
    @action(detail=True, methods=['post'])
    def toggle_like(self, request, pk=None):
        """Toggle like on a note"""
        try:
            note = self.get_object()
            user = request.user
            
            if note.likes.filter(id=user.id).exists():
                note.likes.remove(user)
                liked = False
                message = 'Like removed'
            else:
                note.likes.add(user)
                liked = True
                message = 'Note liked'
            
            return Response({
                'success': True,
                'data': {
                    'liked': liked,
                    'likes_count': note.likes_count
                },
                'message': message
            })
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def export_projects_excel(request):
    """
    Export projects and tasks to Excel with 3 sheets:
    1. Projects - All projects with full details
    2. Tasks - All tasks with full details
    3. Projects with Tasks - Projects with their related tasks
    """
    try:
        # Create workbook
        wb = Workbook()
        
        # Styling
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ============ SHEET 1: PROJECTS ============
        ws_projects = wb.active
        ws_projects.title = "Projets"
        
        # Project headers
        project_headers = [
            "N° Projet", "Nom", "Description", "Statut", "Priorité", "Catégorie",
            "Chef de Projet", "Fonction CP", "Équipe (Nb)", "Membres Équipe",
            "Date Début", "Date Fin", "Date Terminé", "Budget", "Dépensé",
            "Progrès (%)", "Tâches Totales", "Tâches Terminées", "Tags/Filiales",
            "Notes", "En Retard", "Utilisation Budget (%)", "Créé le", "Modifié le"
        ]
        
        # Write project headers
        for col, header in enumerate(project_headers, 1):
            cell = ws_projects.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Get all projects with related data
        projects = Project.objects.all().prefetch_related('team', 'manager', 'tasks')
        
        # Write project data
        for row, project in enumerate(projects, 2):
            team_members = ", ".join([m.full_name or m.username for m in project.team.all()])
            tags_str = ", ".join(project.tags) if project.tags else ""
            
            project_data = [
                project.project_number or "",
                project.name,
                project.description or "",
                project.get_status_display(),
                project.get_priority_display(),
                project.get_category_display(),
                project.manager.full_name if project.manager else "",
                getattr(project.manager, 'position', '') if project.manager else "",
                project.team_count,
                team_members,
                project.start_date.strftime('%d/%m/%Y') if project.start_date else "",
                project.deadline.strftime('%d/%m/%Y') if project.deadline else "",
                project.completed_date.strftime('%d/%m/%Y') if project.completed_date else "",
                f"{project.budget:.2f} €" if project.budget else "0.00 €",
                f"{project.spent:.2f} €" if project.spent else "0.00 €",
                project.progress,
                project.get_tasks_count(),
                project.get_completed_tasks_count(),
                tags_str,
                project.notes or "",
                "Oui" if project.is_overdue else "Non",
                f"{project.budget_utilization:.1f} %" if project.budget > 0 else "0 %",
                project.created_at.strftime('%d/%m/%Y %H:%M') if project.created_at else "",
                project.updated_at.strftime('%d/%m/%Y %H:%M') if project.updated_at else "",
            ]
            
            for col, value in enumerate(project_data, 1):
                cell = ws_projects.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        # Auto-size columns for projects
        for col in range(1, len(project_headers) + 1):
            ws_projects.column_dimensions[get_column_letter(col)].width = 15
        
        # ============ SHEET 2: TASKS ============
        ws_tasks = wb.create_sheet("Tâches")
        
        # Task headers
        task_headers = [
            "N° Tâche", "Titre", "Description", "Projet", "N° Projet", "Statut",
            "Priorité", "Type", "Assigné à", "Rapporteur", "Date Échéance",
            "Date Terminé", "Temps Estimé (h)", "Temps Réel (h)", "Variance (%)",
            "Progrès (%)", "Tags", "Notes", "En Retard", "Créé le", "Modifié le"
        ]
        
        # Write task headers
        for col, header in enumerate(task_headers, 1):
            cell = ws_tasks.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Get all tasks
        tasks = Task.objects.all().select_related('project', 'assignee', 'reporter')
        
        # Write task data
        for row, task in enumerate(tasks, 2):
            task_tags = ", ".join(task.tags) if task.tags else ""
            
            task_data = [
                task.task_number or "",
                task.title,
                task.description or "",
                task.project.name if task.project else "",
                task.project.project_number if task.project else "",
                task.get_status_display(),
                task.get_priority_display(),
                task.get_task_type_display(),
                task.assignee.full_name if task.assignee else "",
                task.reporter.full_name if task.reporter else "",
                task.due_date.strftime('%d/%m/%Y') if task.due_date else "",
                task.completed_date.strftime('%d/%m/%Y') if task.completed_date else "",
                task.estimated_time or 0,
                task.actual_time or 0,
                f"{task.time_variance:.1f} %" if task.estimated_time and task.estimated_time > 0 else "0 %",
                task.progress_percentage,
                task_tags,
                task.notes or "",
                "Oui" if task.is_overdue else "Non",
                task.created_at.strftime('%d/%m/%Y %H:%M') if task.created_at else "",
                task.updated_at.strftime('%d/%m/%Y %H:%M') if task.updated_at else "",
            ]
            
            for col, value in enumerate(task_data, 1):
                cell = ws_tasks.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        # Auto-size columns for tasks
        for col in range(1, len(task_headers) + 1):
            ws_tasks.column_dimensions[get_column_letter(col)].width = 15
        
        # ============ SHEET 3: PROJECTS WITH TASKS ============
        ws_combined = wb.create_sheet("Projets avec Tâches")
        
        # Combined headers
        combined_headers = [
            "Type", "N° Projet", "Nom Projet", "Statut Projet", "Chef de Projet", "Progrès Projet (%)",
            "N° Tâche", "Titre Tâche", "Statut Tâche", "Assigné à", "Échéance Tâche",
            "Progrès Tâche (%)", "Temps Estimé (h)", "Temps Réel (h)"
        ]
        
        # Write combined headers
        for col, header in enumerate(combined_headers, 1):
            cell = ws_combined.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Styling for project rows and task rows
        project_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
        task_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        project_font_bold = Font(bold=True, size=11)
        task_indent = Alignment(horizontal="left", vertical="top", indent=2)
        
        # Write combined data with visual hierarchy
        current_row = 2
        for project in projects:
            project_tasks = project.tasks.all()
            
            # Write project header row
            project_row_data = [
                "PROJET",  # Type indicator
                project.project_number or "",
                project.name,
                project.get_status_display(),
                project.manager.full_name if project.manager else "",
                project.progress,
                "", "", "", "", "", "", "", ""  # Empty task columns
            ]
            
            for col, value in enumerate(project_row_data, 1):
                cell = ws_combined.cell(row=current_row, column=col, value=value)
                cell.border = border
                cell.fill = project_fill
                if col <= 6:  # Project columns
                    cell.font = project_font_bold
                cell.alignment = Alignment(vertical="top", horizontal="left")
            
            current_row += 1
            
            # Write tasks as sub-rows
            if project_tasks.exists():
                for task in project_tasks:
                    task_row_data = [
                        "  → Tâche",  # Indented type indicator
                        "", "", "", "", "",  # Empty project columns
                        task.task_number or "",
                        task.title,
                        task.get_status_display(),
                        task.assignee.full_name if task.assignee else "",
                        task.due_date.strftime('%d/%m/%Y') if task.due_date else "",
                        task.progress_percentage,
                        task.estimated_time or 0,
                        task.actual_time or 0,
                    ]
                    
                    for col, value in enumerate(task_row_data, 1):
                        cell = ws_combined.cell(row=current_row, column=col, value=value)
                        cell.border = border
                        cell.fill = task_fill
                        if col == 1:  # Type column with indent
                            cell.alignment = task_indent
                        else:
                            cell.alignment = Alignment(vertical="top")
                    
                    current_row += 1
            else:
                # Project without tasks - add a note row
                no_task_row = [
                    "  (Aucune tâche)",
                    "", "", "", "", "", "", "", "", "", "", "", "", ""
                ]
                
                for col, value in enumerate(no_task_row, 1):
                    cell = ws_combined.cell(row=current_row, column=col, value=value)
                    cell.border = border
                    cell.fill = task_fill
                    cell.font = Font(italic=True, color="999999")
                    cell.alignment = task_indent
                
                current_row += 1
        
        # Auto-size columns for combined
        ws_combined.column_dimensions['A'].width = 15  # Type column
        for col in range(2, len(combined_headers) + 1):
            ws_combined.column_dimensions[get_column_letter(col)].width = 18
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename=export_projets_{timestamp}.xlsx'
        
        # Save workbook to response
        wb.save(response)
        
        return response
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e),
            'message': 'Erreur lors de l\'export Excel'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
